'''
Created on 30 нояб. 2021 г.

@author: Arthur Stankevich
'''

import openpyxl
# import csv

cells = ['d2', 'd3',
         'D16', 'D17', 'D18', 'D22', 'D23', 'D24', 'D25',
         'd30', 'e30', 'f30', 'g30', 'h30',
         'd33', 'e33', 'f33', 'g33', 'h33',
         'd37', 'e37', 'f37', 'g37', 'h37',
         'd39', 'e39', 'f39', 'g39', 'h39',
         'd19', 'd27',
         'd44', 'd45', 'd46', 'd47', 'd48', 'd49', 'd50',
         'e44', 'e45', 'e46', 'e47', 'e48', 'e49', 'e50',
         'f44', 'f45', 'f46', 'f47', 'f48', 'f49', 'f50',
         'g44', 'g45', 'g46', 'g47', 'g48', 'g49', 'g50',
         'h44', 'h45', 'h46', 'h47', 'h48', 'h49', 'h50',
         'i44', 'i45', 'i46', 'i47', 'i48', 'i49', 'i50',
         'd54', 'd57', 'e54', 'e57', 'f54', 'f57', 'g54', 'g57', 'h54', 'h57', 'i54', 'i57',
         'j54', 'j57', 'k54', 'k57', 'l54', 'l57', 'm54', 'm57', 'n54', 'n57', 'o54', 'o57'
         ]

def read_bankId():
    # with open('banks_list.csv') as file:
    #     list = csv.reader(file)
    # print(list)
    file = openpyxl.load_workbook('banks_list.xlsx')
    sheet = file['list']
    i = 2
    banks = {}
    while True:
        bank = sheet.cell(row=i, column=2).value
        # print(bank)
        if bank is None:
            break
        bank_id = sheet.cell(row=i, column=3).value
        # print(bank_id)
        banks[bank] = bank_id
        # print(banks)
        i += 1
    return (banks)


def save_to_excel(filename, values):
    temp_file = openpyxl.load_workbook('data_template.xlsx')
    temp_sheet = temp_file.active
    for c, v in zip(cells, values):
        cell = temp_sheet[c]
        cell.value = v
    temp_file.save(filename+'.xlsx')

# read_bankId()